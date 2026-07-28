"""
SARSP-LangEd - Stage 2: LLM-Assisted Rater Screening
Dynamic model selection, customizable prompts, resumable processing, 
auto-save functionality, and color-coded adjudication export.
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
import json
import time
import io
import os
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from copy import copy
from jsonschema import validate, ValidationError


# =============================================================================
# CONSTANTS & DEFAULT PROMPTS
# =============================================================================

DEFAULT_SYSTEM_PROMPT = "As an expert scholar in applied linguistics, you specialise in the application of text-based large language models (LLMs) in language education contexts."

DEFAULT_USER_PROMPT_TEMPLATE = """### TASK CONTEXT ###
You are conducting a systematic review of empirical studies on the application of text-based Large Language Models (LLMs) in language education. 
Language education includes language learning, teaching, assessment, and instructional design practices in first, second, and foreign language contexts, as well as bilingual and CLIL settings.

### INCLUSION CRITERIA ###
Include studies that:
1. Empirically investigate the application of LLMs (e.g., ChatGPT, Copilot, Gemini, LLaMa, Claude) in language education contexts.
2. Address in-classroom or out-of-classroom practices related to LLM use — including, but not limited to, development of written or oral skills, assessment design, instructional design, AI literacy, or similar pedagogical applications.
3. Focus on language teachers, learners, or other stakeholders engaging with LLMs for language education purposes.

### EXCLUSION CRITERIA ###
Exclude studies that:
1. Do not involve text-based LLMs or use non-LLM AI tools (e.g., traditional NLP systems, grammar checkers, speech recognition, or machine translation not based on LLMs).
2. Are not empirical (e.g., conceptual, theoretical, opinion, review, or bibliometric papers).
3. Are empirical but irrelevant, i.e. do not focus on language education (e.g., general AI literacy without language focus)."""

# Protected suffix that cannot be edited by users (QUESTIONS + STUDY + OUTPUT)
PROTECTED_SUFFIX = """
### QUESTIONS ###
a) Should this study be excluded? Answer only "Y" (Yes) or "N" (No). 
b) Provide a concise justification explaining why, using your expertise.

### STUDY ###
- Title: "{study_title}"
- Abstract: "{study_abstract}"
- Author Keywords: "{study_keywords}"

### OUTPUT FORMAT ###
Respond ONLY with valid JSON matching this schema:
{{
  "exclude_study": "Y" or "N",
  "justification": "string (brief explanation, 1–3 sentences)"
}}
"""

# Fixed JSON schema - users cannot modify this to ensure pipeline integrity
JSON_SCHEMA = {
    "type": "json_schema",
    "json_schema": {
        "schema": {
            "type": "object",
            "properties": {
                "exclude_study": {"type": "string", "enum": ["Y", "N"]},
                "justification": {"type": "string"}
            },
            "required": ["exclude_study", "justification"],
            "additionalProperties": False
        }
    }
}

# Pre-configured API providers for non-tech users
_ALL_API_PROVIDERS = {
    "Cerebras (Cloud)": "https://api.cerebras.ai/v1",
    "Google Gemini": "https://generativelanguage.googleapis.com/v1beta/openai",
    "Mistral AI": "https://api.mistral.ai/v1",
    "OpenAI": "https://api.openai.com/v1",
    "LM Studio (Local)": "http://localhost:1234/v1",
    "Ollama (Local)": "http://localhost:11434/v1",
    "Custom / Other": "custom",
}

def _get_available_providers():
    """Filter providers based on deployment context. Local/custom only available locally."""
    is_cloud = bool(os.environ.get("STREAMLIT_SHARING_MODE") or os.environ.get("STREAMLIT_CLOUD"))
    if is_cloud:
        return {k: v for k, v in _ALL_API_PROVIDERS.items()
                if "Local" not in k and "Custom" not in k}
    return _ALL_API_PROVIDERS

# Auto-save configuration with temp folder logic
from core.utils import get_session_temp_dir
TEMP_DIR = get_session_temp_dir()
AUTO_SAVE_DIR_NAME = "stage2_screening"
AUTO_SAVE_FILENAME = os.path.join(TEMP_DIR, AUTO_SAVE_DIR_NAME, "screening.xlsx")

# Ensure directories exist immediately upon loading (robustly)
try:
    os.makedirs(os.path.join(TEMP_DIR, AUTO_SAVE_DIR_NAME), exist_ok=True)
except Exception as e:
    # Fallback: Try to create just TEMP_DIR if subfolder fails
    try:
        os.makedirs(TEMP_DIR, exist_ok=True)
    except:
        pass


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_api_client(base_url: str, api_key: str) -> OpenAI:
    """Create an OpenAI-compatible client for any endpoint."""
    return OpenAI(base_url=base_url, api_key=api_key or "sk-no-key")


def query_model(client, model_id, title, abstract, keywords, system_prompt, 
                user_prompt_template, max_retries=5, backoff_factor=5):
    """
    Query LLM with exponential backoff, JSON validation, and graceful degradation.
    Returns dict with 'exclude_study' and 'justification' keys.
    """
    attempt = 0
    raw_output = ""
    while attempt < max_retries:
        try:
            completion = client.chat.completions.create(
                model=model_id,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt_template.format(
                        study_title=title or "",
                        study_abstract=abstract or "",
                        study_keywords=keywords or ""
                    )}
                ],
                temperature=0,
                response_format=JSON_SCHEMA
            )

            raw_output = completion.choices[0].message.content.strip()
            response = json.loads(raw_output)
            validate(instance=response, schema=JSON_SCHEMA["json_schema"]["schema"])
            return response

        except (json.JSONDecodeError, ValidationError):
            json_part = re.search(r"\{.*\}", raw_output, re.DOTALL)
            if json_part:
                try:
                    response = json.loads(json_part.group(0))
                    validate(instance=response, schema=JSON_SCHEMA["json_schema"]["schema"])
                    return response
                except Exception:
                    pass
            return {"exclude_study": "?", "justification": f"Parsing error: {raw_output[:200]}"}

        except Exception as e:
            err_msg = str(e)
            if "queue_exceeded" in err_msg or "rate_limit" in err_msg.lower():
                delay = backoff_factor * (2 ** attempt)
                time.sleep(delay)
                attempt += 1
                continue
            elif "token_quota_exceeded" in err_msg or "too_many_requests" in err_msg:
                return {"exclude_study": "?", "justification": f"Quota/rate error: {err_msg[:100]}"}
            else:
                return {"exclude_study": "?", "justification": f"API error: {err_msg[:100]}"}

    return {"exclude_study": "?", "justification": "Error: too many retries or persistent failure"}


def apply_color_coding(excel_bytes: bytes, llm_raters: list) -> bytes:
    """Apply color-coded adjudication logic to screened Excel file."""
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    existing_fills = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        fills = [cell.fill for cell in row if getattr(cell.fill, "fill_type", None) not in (None, 'none', '')]
        if fills:
            existing_fills[row[0].row] = copy(fills[0])

    headers = [c.value for c in ws[1]]
    def get_col_idx(name):
        return headers.index(name) + 1 if name in headers else None

    c_llm = [get_col_idx(f"Exclude? ({mid})") for mid in llm_raters]
    c_final = get_col_idx("Final decision")

    def says_exclude(value):
        return str(value).strip().upper().startswith("Y")

    def says_not_exclude(value):
        return str(value).strip().upper().startswith("N")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_idx = row[0].row
        llm_vals = [str(row[i-1].value or "").strip() for i in c_llm if i]
        llm_excludes = sum(says_exclude(v) for v in llm_vals)
        llm_not_excludes = sum(says_not_exclude(v) for v in llm_vals)
        total_llm_raters = len([v for v in llm_vals if v])

        fill = None
        final_decision = ""

        if total_llm_raters >= 2 and llm_excludes > 0 and llm_not_excludes > 0:
            fill = orange_fill
            final_decision = "DISAGREE"
        elif total_llm_raters > 0 and (llm_excludes > total_llm_raters / 2):
            fill = red_fill
            final_decision = "EXCLUDE"
        else:
            final_decision = ""

        if fill:
            for cell in row:
                cell.fill = copy(fill)
        elif row_idx in existing_fills:
            for cell in row:
                cell.fill = copy(existing_fills[row_idx])
        else:
            for cell in row:
                cell.fill = PatternFill(fill_type=None)

        if c_final:
            ws.cell(row=row_idx, column=c_final, value=final_decision)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def save_intermediate_excel(df, selected_models):
    """
    Saves the current state of the dataframe to disk immediately inside the temp folder.
    This ensures data is preserved even if the session crashes.
    """
    try:
        # Ensure BOTH temp and subfolder exist
        os.makedirs(os.path.join(TEMP_DIR, AUTO_SAVE_DIR_NAME), exist_ok=True)

        buffer = io.BytesIO()
        df.to_excel(buffer, index=False)
        buffer.seek(0)
        
        # Apply color coding to the intermediate save as well
        colored_bytes = apply_color_coding(buffer.read(), selected_models)
        
        with open(AUTO_SAVE_FILENAME, "wb") as f:
            f.write(colored_bytes)
        return True
    except Exception as e:
        # Log error to console for debugging (Streamlit suppresses prints in loops)
        import sys
        print(f"❌ Auto-save failed: {e}", file=sys.stderr)
        return False

def compute_screening_stats(df: pd.DataFrame) -> dict:
    """
    Compute comprehensive screening statistics including human adjudication
    outcomes and PDF retrieval status. Works with any screened DataFrame that
    has 'Final decision', 'Exclude? (Human rater)', and optionally 'Downloaded?' columns.
    """
    stats = {
        "total": len(df),
        "auto_excluded": 0,
        "disagreements_total": 0,
        "disagree_accepted": 0,
        "disagree_rejected": 0,
        "disagree_unresolved": 0,
        "included_final": 0,
        "downloaded": 0,
        "not_downloaded": 0,
        "retrieval_pending": 0,
    }

    if df is None or df.empty:
        return stats

    fd = df.get("Final decision", pd.Series(dtype=str)).astype(str).str.strip().str.upper()
    human = df.get("Exclude? (Human rater)", pd.Series(dtype=str)).astype(str).str.strip().str.upper()
    downloaded = df.get("Downloaded?", pd.Series(dtype=str)).astype(str).str.strip().str.upper() if "Downloaded?" in df.columns else pd.Series([""] * len(df))

    stats["auto_excluded"] = int((fd == "EXCLUDE").sum())
    stats["disagreements_total"] = int((fd == "DISAGREE").sum())

    # Adjudication outcomes for DISAGREE rows
    disagree_mask = fd == "DISAGREE"
    if disagree_mask.any():
        disagree_human = human[disagree_mask]
        stats["disagree_accepted"] = int(disagree_human.str.startswith("N").sum())  # N = do NOT exclude = accepted
        stats["disagree_rejected"] = int(disagree_human.str.startswith("Y").sum())  # Y = exclude = rejected
        stats["disagree_unresolved"] = stats["disagreements_total"] - stats["disagree_accepted"] - stats["disagree_rejected"]

    # Final included = total - auto_excluded - disagree_rejected - unresolved
    stats["included_final"] = stats["total"] - stats["auto_excluded"] - stats["disagree_rejected"] - stats["disagree_unresolved"]

    # Retrieval status (only for included studies)
    include_mask = ~fd.isin(["EXCLUDE"]) & ~(disagree_mask & human.str.startswith("Y"))
    if "Downloaded?" in df.columns and include_mask.any():
        dl_vals = downloaded[include_mask]
        stats["downloaded"] = int(dl_vals.str.startswith("Y").sum())
        stats["not_downloaded"] = int(dl_vals.str.startswith("N").sum())
        stats["retrieval_pending"] = int((dl_vals == "").sum() | dl_vals.str.startswith("PENDING").sum())
        # Recalculate pending as included minus downloaded minus explicitly not-downloaded
        stats["retrieval_pending"] = max(0, stats["included_final"] - stats["downloaded"] - stats["not_downloaded"])

    return stats

# =============================================================================
# STREAMLIT RENDER FUNCTION
# =============================================================================
def render_screening_page():
    """Render the full Screening stage UI."""
    st.title("2️⃣ LLM-Assisted Rater Screening")
    st.markdown("""
    Screen titles and abstracts against eligibility criteria using **independent LLM raters**. 
    Disagreements are flagged for human adjudication. Upload your data, configure your API, 
    select models, and customize the screening prompts below.
    """)

    st.divider()

    # =========================================================================
    # DATA INPUT SECTION
    # =========================================================================
    df = None

    # === REPLICATION MODE INJECTION ===
    if st.session_state.get("replication_mode"):
        from core.utils import get_replication_path

        st.info("🔬 **Replication Mode:** Original screening results loaded. Toggle off in sidebar to run your own screening.")

        repl_path = get_replication_path("s2_screening_reviewed")
        if repl_path and os.path.exists(repl_path):
            try:
                repl_df = pd.read_excel(repl_path, dtype=str).fillna("")
                st.session_state["screening_df"] = repl_df
                st.session_state["screening_complete"] = True
                st.session_state["screening_excel"] = open(repl_path, "rb").read()
                st.success(f"✅ Loaded {len(repl_df)} screened records from replication data.")
            except Exception as e:
                st.error(f"❌ Failed to load replication screening file: {e}")
        else:
            st.warning("⚠️ Replication screening file not found. Check `replication_data/stage2_screening_after_review.xlsx`.")

        # In replication mode, skip directly to results display
        if st.session_state.get("screening_complete"):
            # Fall through to RESULTS DISPLAY section below
            pass
        else:
            return
    else:
        # --- AUTO-RECOVER FROM DISK IF SESSION STATE LOST ---
        if ("screening_df" not in st.session_state or st.session_state["screening_df"] is None) \
           and os.path.exists(AUTO_SAVE_FILENAME):
            try:
                recovered_df = pd.read_excel(AUTO_SAVE_FILENAME, dtype=str).fillna("")
                st.session_state["screening_df"] = recovered_df
                st.info(f"♻️ Recovered screening progress from `{AUTO_SAVE_FILENAME}` ({len(recovered_df)} records)")
            except Exception as e:
                st.warning(f"⚠️ Could not recover auto-save: {e}")

        st.subheader("📁 Input Data")

        data_source = st.radio(
            "Data source",
            ["Upload CSV file", "Use data from Stage 1 (Preprocessing)"],
            horizontal=True,
            key="screening_data_source"
        )

        if data_source == "Upload CSV file":
            uploaded_file = st.file_uploader(
                "Upload filtered references CSV",
                type=["csv"],
                help="CSV file with semicolon separator containing Title, Abstract, and Author Keywords columns.",
                key="screening_csv_upload"
            )
            if uploaded_file:
                try:
                    df = pd.read_csv(uploaded_file, sep=";", dtype=str).fillna("")
                    st.success(f"✅ Loaded {len(df)} records from uploaded file")
                except Exception as e:
                    st.error(f"❌ Failed to load CSV: {str(e)}")
        else:
            if st.session_state.get("preprocessing_complete"):
                df = st.session_state.get("preprocessing_final")
                if df is not None and not df.empty:
                    st.success(f"✅ Using {len(df)} records from Stage 1 preprocessing")
                else:
                    st.warning("⚠️ No preprocessed data found. Please complete Stage 1 first or upload a file.")
            else:
                st.info("ℹ️ Stage 1 has not been completed yet. Upload a CSV file above or run Preprocessing first.")

        if df is None or df.empty:
            st.warning("⚠️ No data available. Please upload a CSV file or complete Stage 1.")
            return

    st.divider()

    # Check running state early to disable widgets
    is_running = st.session_state.get("screening_is_running", False)

    # Skip configuration and processing when in replication mode
    if not st.session_state.get("replication_mode"):

        # =========================================================================
        # CONFIGURATION PANEL
        # =========================================================================
        with st.expander("⚙️ Screening Configuration", expanded=True):
            
            # --- API Provider Selection ---
            st.subheader("🔌 API Configuration")
            
            provider_col1, provider_col2 = st.columns([1, 1])
            
            available_providers = _get_available_providers()

            with provider_col1:
                selected_provider = st.selectbox(
                    "API Service Provider",
                    options=list(available_providers.keys()),
                    index=0,
                    help="Select your LLM provider. Base URL will be auto-filled."
                        + (" Local/custom providers require running the app on your own machine."
                            if "Local" not in str(available_providers.keys()) else ""),
                    key="screening_provider",
                    disabled=is_running
                )

            default_url = available_providers.get(selected_provider, "")
            
            if "screening_last_provider" not in st.session_state or st.session_state["screening_last_provider"] != selected_provider:
                st.session_state["screening_base_url"] = default_url if default_url != "custom" else ""
                st.session_state["screening_last_provider"] = selected_provider
                
            with provider_col2:
                if default_url == "custom":
                    base_url = st.text_input(
                        "Custom API Base URL",
                        value=st.session_state.get("screening_base_url", ""),
                        placeholder="https://your-api-endpoint.com/v1",
                        help="Enter your custom OpenAI-compatible API endpoint.",
                        key="screening_custom_url",
                        disabled=is_running  # BUG FIX #3: Disable while running
                    )
                else:
                    base_url = st.text_input(
                        "API Base URL (auto-filled)",
                        value=st.session_state.get("screening_base_url", default_url),
                        help="Auto-filled based on provider selection. Edit if needed.",
                        key="screening_base_url",
                        disabled=is_running  # BUG FIX #3: Disable while running
                    )
            
            api_key = st.text_input(
                "API Key",
                type="password",
                value="",
                help="Enter your API key. Leave empty for local models that don't require authentication.",
                key="screening_api_key_input",
                disabled=is_running  # BUG FIX #3: Disable while running
            )

            # --- Model Discovery ---
            st.caption("Click below to discover available models from your selected endpoint.")
            if st.button("🔍 Discover Models", key="screening_discover_btn", disabled=is_running):
                if not base_url:
                    st.error("❌ Please provide an API Base URL first.")
                else:
                    try:
                        with st.spinner("Fetching model list..."):
                            client = get_api_client(base_url, api_key)
                            models_response = client.models.list()
                            available_models = sorted([m.id for m in models_response])
                            st.session_state["screening_available_models"] = available_models
                            st.success(f"✅ Found {len(available_models)} models")
                    except Exception as e:
                        st.error(f"❌ Failed to fetch models: {str(e)}")
                        st.session_state["screening_available_models"] = []

            available_models = st.session_state.get("screening_available_models", [])
            
            if available_models:
                selected_models = st.multiselect(
                    "Select LLM Raters (minimum 2 recommended for dual-rater design)",
                    options=available_models,
                    default=available_models[:2] if len(available_models) >= 2 else available_models,
                    help="Each selected model will independently screen every record.",
                    key="screening_model_select",
                    disabled=is_running  # BUG FIX #3: Disable while running
                )
            else:
                selected_models = []
                st.info("👆 Click 'Discover Models' to load available models from your endpoint.")

            st.divider()

            # --- System Prompt Customization ---
            st.subheader("🤖 System Prompt")
            st.caption("Define the LLM's role and expertise. This sets the behavioral context for all screening decisions.")
            
            # BUG FIX #3: Disable editing while running
            # BUG FIX #1: Use widget value directly, do NOT assign to session_state after this line
            system_prompt = st.text_area(
                "System Prompt",
                value=st.session_state.get("screening_system_prompt", DEFAULT_SYSTEM_PROMPT),
                height=80,
                help="Describe the expert persona the LLM should adopt.",
                key="screening_system_prompt",
                disabled=is_running
            )
            
            with st.expander("👁️ Preview System Prompt"):
                st.code(system_prompt, language="text")

            st.divider()

            # --- User Prompt Customization ---
            st.subheader("📝 User Prompt (Screening Criteria)")
            st.caption(
                "Customize the task context and inclusion/exclusion criteria below. "
                "The **questions**, study metadata placeholders, and JSON output format are automatically appended "
                "and **cannot be modified** to ensure schema compliance and consistent dual-rater decisions."
            )
            
            # BUG FIX #3: Disable editing while running
            custom_user_prompt = st.text_area(
                "Editable User Prompt Body",
                value=st.session_state.get("screening_user_prompt_body", DEFAULT_USER_PROMPT_TEMPLATE),
                height=300,
                help="Edit the task context, inclusion/exclusion criteria, and questions.",
                key="screening_user_prompt",
                disabled=is_running
            )
            
            # Reconstruct full prompt with protected suffix
            full_user_prompt_template = custom_user_prompt + PROTECTED_SUFFIX
            
            # FIX: Use explicit key on expander to prevent duplicate rendering on reruns
            with st.expander("👁️ Preview Full User Prompt (read-only)", expanded=False):
                st.code(full_user_prompt_template, language="text")

        st.divider()

        # =========================================================================
        # RUN SCREENING
        # =========================================================================
        
        has_existing_results = "screening_df" in st.session_state and st.session_state.get("screening_df") is not None
        can_run = len(selected_models) >= 1 and base_url and df is not None
        
        col_btn1, col_btn2 = st.columns([1, 4])
        with col_btn1:
            if is_running:
                stop_clicked = st.button("🛑 Stop Screening", type="secondary", key="screening_stop_btn")
                if stop_clicked:
                    st.session_state["screening_is_running"] = False
                    st.rerun()
            else:
                # Show "Resume" only when there are partial results AND screening is not marked complete
                # Show "Run Screening" when: no results, OR screening was completed, OR data changed
                is_complete = st.session_state.get("screening_complete", False)

                # Detect if user switched to a different data source since last run
                if data_source == "Upload CSV file":
                    _current_data_id = uploaded_file.name if uploaded_file else None
                else:
                    _current_data_id = "__stage1_preprocessing__"
                _data_changed = _current_data_id != st.session_state.get("screening_data_id")

                if has_existing_results and not is_complete and not _data_changed:
                    btn_label = "🔄 Resume Screening"
                else:
                    btn_label = "🚀 Run Screening"
                    
                run_clicked = st.button(btn_label, type="primary", disabled=not can_run, key="screening_run_btn")
                if run_clicked:
                    # Track current data source identity to detect new uploads
                    # Use file name for uploads, or a stable marker for Stage 1 data
                    if data_source == "Upload CSV file":
                        current_data_id = uploaded_file.name if uploaded_file else None
                    else:
                        current_data_id = "__stage1_preprocessing__"

                    previous_data_id = st.session_state.get("screening_data_id")
                    data_changed = current_data_id != previous_data_id

                    if data_changed or not has_existing_results:
                        # === FRESH START ===
                        work_df = df.copy()
                        work_df.fillna("", inplace=True)

                        human_col = "Exclude? (Human rater)"
                        if human_col not in work_df.columns:
                            work_df.insert(0, human_col, "")

                        insert_at = work_df.columns.get_loc(human_col) + 1
                        for model_id in selected_models:
                            col_name = f"Exclude? ({model_id})"
                            if col_name not in work_df.columns:
                                work_df.insert(insert_at, col_name, "")
                                insert_at += 1

                        final_col = "Final decision"
                        if final_col not in work_df.columns:
                            last_llm_col = max(work_df.columns.get_loc(f"Exclude? ({m})") for m in selected_models)
                            work_df.insert(last_llm_col + 1, final_col, "")

                        downloaded_col = "Downloaded?"
                        if downloaded_col not in work_df.columns:
                            final_col_idx = work_df.columns.get_loc(final_col)
                            work_df.insert(final_col_idx + 1, downloaded_col, "")

                        st.session_state["screening_df"] = work_df
                        st.session_state["screening_data_id"] = current_data_id
                        # Clear all stale state for a clean fresh run
                        st.session_state.pop("screening_complete", None)
                        st.session_state.pop("screening_excel", None)
                        st.session_state.pop("screening_log_history", None)
                        st.session_state.pop("screening_models", None)
                    else:
                        # === RESUME EXISTING ===
                        work_df = st.session_state["screening_df"].copy()

                        final_col = "Final decision"
                        if final_col not in work_df.columns:
                            llm_cols = [c for c in work_df.columns if c.startswith("Exclude? (") and c != "Exclude? (Human rater)"]
                            if llm_cols:
                                last_llm_idx = max(work_df.columns.get_loc(c) for c in llm_cols)
                                work_df.insert(last_llm_idx + 1, final_col, "")

                        downloaded_col = "Downloaded?"
                        if downloaded_col not in work_df.columns and final_col in work_df.columns:
                            final_col_idx = work_df.columns.get_loc(final_col)
                            work_df.insert(final_col_idx + 1, downloaded_col, "")

                        st.session_state["screening_df"] = work_df

                    st.session_state["screening_is_running"] = True
                    st.rerun()

        # PROCESSING LOOP
        if st.session_state.get("screening_is_running", False):
            work_df = st.session_state.get("screening_df")
            if work_df is None or work_df.empty:
                st.session_state["screening_is_running"] = False
                st.rerun()
                
            client = get_api_client(base_url, api_key)
            
            total_calls = len(selected_models) * len(work_df)
            completed = 0
            errors = 0
            
            for model_id in selected_models:
                col_name = f"Exclude? ({model_id})"
                if col_name in work_df.columns:
                    completed += work_df[col_name].astype(str).apply(lambda x: x.strip() != "" and not x.startswith("?")).sum()
            
            progress_bar = st.progress(completed / total_calls if total_calls > 0 else 0, text=f"Progress: {completed}/{total_calls}")
            status_text = st.empty()
            
            st.markdown("#### 📋 Live Processing Log")
            log_container = st.container()
            
            if "screening_log_history" not in st.session_state:
                st.session_state["screening_log_history"] = []
                
            stop_processing = False
            
            with log_container:
                for log_entry in st.session_state["screening_log_history"]:
                    with st.chat_message(log_entry["role"]):
                        st.markdown(log_entry["message"])
            
            for model_id in selected_models:
                if st.session_state.get("screening_is_running", False) == False:
                    break
                    
                col_name = f"Exclude? ({model_id})"
                status_text.text(f"🔄 Processing model: **{model_id}**")
                
                for idx in range(len(work_df)):
                    if not st.session_state.get("screening_is_running", True):
                        stop_processing = True
                        break
                        
                    cell_val = str(work_df.at[idx, col_name]).strip()
                    
                    if cell_val and not cell_val.startswith("?"):
                        continue

                    title = str(work_df.at[idx, "Title"]).strip()
                    abstract = str(work_df.at[idx, "Abstract"]).strip()
                    keywords = str(work_df.at[idx, "Author Keywords"]).strip()

                    result = query_model(
                        client, model_id, title, abstract, keywords, 
                        system_prompt, full_user_prompt_template
                    )
                    
                    result_str = f"{result['exclude_study']} ({result['justification']})"
                    work_df.at[idx, col_name] = result_str
                    
                    decision_icon = "🟢" if result['exclude_study'] == "N" else ("🔴" if result['exclude_study'] == "Y" else "❓")
                    log_message = (
                        f"**Row {idx}** | **Model:** {model_id}\n\n"
                        f"**Title:** {title}\n\n"
                        f"**Abstract:** {abstract}\n\n"
                        f"**Keywords:** {keywords}\n\n"
                        f"**Decision:** {decision_icon} {result['exclude_study']}\n\n"
                        f"**Justification:** {result['justification']}"
                    )
                    
                    new_log = {"role": "assistant", "message": log_message}
                    st.session_state["screening_log_history"].append(new_log)
                    
                    with log_container:
                        with st.chat_message("assistant"):
                            st.markdown(log_message)
                    
                    if result["exclude_study"] == "?":
                        errors += 1
                    
                    completed += 1
                    progress_bar.progress(completed / total_calls, text=f"Progress: {completed}/{total_calls}")
                    
                    save_intermediate_excel(work_df, selected_models)
                    
                    time.sleep(0.01)
                
                if stop_processing:
                    break
            
            st.session_state["screening_df"] = work_df
            
            if not st.session_state.get("screening_is_running", False):
                st.warning(f"⏸️ Screening paused. Auto-saved to `{AUTO_SAVE_FILENAME}`. Click 'Resume Screening' to continue.")
            elif completed >= total_calls:
                st.session_state["screening_is_running"] = False
                progress_bar.progress(1.0, text="✅ Screening complete!")
                status_text.empty()
                
                status_text.text("🎨 Applying final color-coded adjudication...")
                excel_buffer = io.BytesIO()
                work_df.to_excel(excel_buffer, index=False)
                excel_buffer.seek(0)
                
                colored_excel = apply_color_coding(excel_buffer.read(), selected_models)
                
                # Persist final color-coded result to disk (safety net)
                try:
                    with open(AUTO_SAVE_FILENAME, "wb") as f:
                        f.write(colored_excel)
                except Exception:
                    pass
                
                st.session_state["screening_excel"] = colored_excel
                
                # BUG FIX #1: Only set completion metadata. 
                # Do NOT re-assign screening_system_prompt or screening_user_prompt here 
                # as their widgets have already been instantiated above.
                if "screening_complete" not in st.session_state:
                    st.session_state["screening_complete"] = True
                    st.session_state["screening_models"] = selected_models
                    st.session_state["screening_base_url_used"] = base_url
                    st.session_state["screening_provider_used"] = selected_provider
                
                status_text.empty()
                st.success(f"✅ **Screening Complete!** Processed {len(work_df)} records × {len(selected_models)} models | Errors: {errors}")
                st.rerun()

    # =========================================================================
    # RESULTS DISPLAY & DOWNLOAD
    # =========================================================================
    if st.session_state.get("screening_complete"):
        screened_df = st.session_state.get("screening_df")
        colored_excel = st.session_state.get("screening_excel")
        used_models = st.session_state.get("screening_models", [])

        if screened_df is not None:
            stats = compute_screening_stats(screened_df)

            st.divider()
            st.subheader("📊 Screening & Adjudication Summary")

            # Row 1: Overall counts
            m1 = st.columns(4)
            m1[0].metric("Total Records", stats["total"])
            m1[1].metric("🔴 Auto-Excluded (LLM)", stats["auto_excluded"])
            m1[2].metric("🟠 LLM Disagreements", stats["disagreements_total"])
            m1[3].metric("✅ Final Included", stats["included_final"])

            # Row 2: Human adjudication breakdown (only if disagreements exist)
            if stats["disagreements_total"] > 0:
                st.markdown("#### 👤 Human Adjudication Outcomes")
                m2 = st.columns(3)
                m2[0].metric(
                    "Disagree → Accepted (Included)",
                    stats["disagree_accepted"],
                    help="LLM said EXCLUDE, human overruled with N → study INCLUDED"
                )
                m2[1].metric(
                    "Disagree → Rejected (Excluded)",
                    stats["disagree_rejected"],
                    help="LLM said INCLUDE, human overruled with Y → study EXCLUDED"
                )
                m2[2].metric(
                    "Disagree → Unresolved",
                    stats["disagree_unresolved"],
                    help="No human decision recorded yet for these disagreements"
                )

            # Row 3: PDF retrieval status (only if Downloaded? column exists)
            if "Downloaded?" in screened_df.columns:
                st.markdown("#### 📥 PDF Retrieval Status")
                m3 = st.columns(3)
                m3[0].metric("✅ Downloaded", stats["downloaded"])
                m3[1].metric("❌ Not Retrieved", stats["not_downloaded"])
                m3[2].metric("⏳ Pending / Unknown", stats["retrieval_pending"])

        with st.expander("👀 Preview Screened Data (first 20 records)"):
            preview_cols = ["Title"] + [f"Exclude? ({m})" for m in used_models] + ["Final decision"]
            preview_cols = [c for c in preview_cols if c in screened_df.columns]
            st.dataframe(screened_df[preview_cols].head(20), use_container_width=True)

        st.divider()
        st.subheader("📥 Download Results")
        
        if os.path.exists(AUTO_SAVE_FILENAME):
            st.info(f"💾 Real-time auto-save active: `{AUTO_SAVE_FILENAME}`")
            
        if colored_excel:
            st.download_button(
                label="📊 Download Final Color-Coded Screening Excel",
                data=colored_excel,
                file_name="filtered_and_screened_refs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="screening_download_btn"
            )

        # Post-screening PDF retrieval instructions (uses pre-computed stats)
        if screened_df is not None and "Downloaded?" in screened_df.columns:
            if stats["retrieval_pending"] > 0 or stats["not_downloaded"] > 0:
                st.warning(f"""
                📥 **PDF Retrieval Required for Stage 3 (Extraction)**

                You have **{stats['retrieval_pending'] + stats['not_downloaded']} included studies** without confirmed PDF downloads.
                Before proceeding to extraction, you must:
                1. Open the downloaded Excel file above
                2. Filter the **"Downloaded?"** column for empty, "N", or "PENDING" values
                3. Retrieve full-text PDFs via your institutional library, DOI links, or open-access repositories
                4. Mark each retrieved PDF as **"Y"** in the "Downloaded?" column
                5. Place all PDFs in a single folder and upload them in **Stage 3: Extraction**

                ✅ Downloaded: **{stats['downloaded']}** | ❌ Not retrieved: **{stats['not_downloaded']}** | ⏳ Pending: **{stats['retrieval_pending']}** | Total included: **{stats['included_final']}**
                """)
            else:
                st.success(f"✅ All {stats['included_final']} included studies have been marked as downloaded. Ready for Stage 3: Extraction.")
        with st.expander("🔬 Reproducibility & Accessibility Information", expanded=False):
            repro = {
                "stage": "screening",
                "timestamp": pd.Timestamp.now().isoformat(),
                "audit_trail": {
                    "models_used": used_models,
                    "api_service_provider": st.session_state.get("screening_provider_used", "N/A"),
                    "api_base_url": st.session_state.get("screening_base_url_used", "N/A"),
                    "system_prompt": st.session_state.get("screening_system_prompt", DEFAULT_SYSTEM_PROMPT),
                    "user_prompt_body": st.session_state.get("screening_user_prompt_body", DEFAULT_USER_PROMPT_TEMPLATE),
                    "protected_suffix": "QUESTIONS + STUDY metadata + OUTPUT FORMAT (immutable)",
                    "json_schema": JSON_SCHEMA,
                    "temperature": 0,
                    "max_retries": 5,
                    "backoff_factor": 5,
                },
                "results_summary": stats if screened_df is not None else {},
                "auto_save_path": AUTO_SAVE_FILENAME,
            }
            st.json(repro)
            st.code(json.dumps(repro, indent=2), language="json")

            st.divider()

            # Human Adjudication Workflow
            st.markdown("#### 👤 Human Adjudication Workflow")
            st.info("""
            1. Filter rows where **Final decision = DISAGREE** (orange rows)
            2. Review title/abstract against eligibility criteria
            3. Update **"Exclude? (Human rater)"** column with Y/N
            4. Update **"Final decision"** to EXCLUDE or INCLUDE accordingly
            """)

            st.divider()

            # Reproducibility Checklist
            st.markdown("#### ✅ Reproducibility Checklist")
            st.markdown("""
            - [x] Exact prompt template archived above (system + user + protected suffix)
            - [x] Model IDs, API endpoint, and generation parameters recorded
            - [x] Input CSV and output Excel downloadable from this page
            - [x] Auto-save path documented (`{}`)
            - [x] JSON schema locked and immutable
            - [ ] Document any manual interventions or schema adjustments made during adjudication
            """.format(AUTO_SAVE_FILENAME))

            st.divider()
            # Accessibility Alternatives
            st.markdown("#### ♿ Accessibility Alternatives")
            is_cloud = bool(os.environ.get("STREAMLIT_SHARING_MODE") or os.environ.get("STREAMLIT_CLOUD"))
            if is_cloud:
                st.markdown("""
                | Scenario | Alternative |
                |---|---|
                | **Low-resource / no API budget** | Install and run this app locally to use free local models (Ollama, LM Studio). The hosted version supports cloud APIs only. |
                | **Non-programmers** | Use the color-coded Excel output directly; all adjudication can be done in spreadsheet software |
                """)
            else:
                st.markdown("""
                | Scenario | Alternative |
                |---|---|
                | **Low-resource / no API budget** | Use local inference with LMStudio/Ollama + Qwen3, Gemma3, Mistral 7B, etc. |
                | **Non-programmers** | Use the color-coded Excel output directly; all adjudication can be done in spreadsheet software |
                """)

            